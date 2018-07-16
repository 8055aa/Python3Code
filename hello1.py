#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Filename:str_format.py
"I am : doestr.__doc__"
import imp
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import animation
from urllib import request
from urllib.request import urlopen
from bs4 import BeautifulSoup
import time
import requests
from bs4 import BeautifulSoup
from multiprocessing import Pool
from openpyxl import Workbook
import json

message = "After editing"
url = "https://movie.douban.com/"

def getTime():
    day = time.strftime("%Y-%m-%d", time.localtime())
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    return day,now

pageNum = 1  # 用来计数，爬取了多少条目的
wb = Workbook()
ws = wb.active
ws.title = '智联招聘'
ws.cell(row=1, column=1).value = '岗位'
ws.cell(row=1, column=2).value = '地址'
ws.cell(row=1, column=3).value = '发表天数'
ws.cell(row=1, column=4).value = '薪资'


def get_zhaopin(page):
    global pageNum
    url = 'http://sou.zhaopin.com/jobs/searchresult.ashx?jl=全国&kw=android&p={0}&kt=3'.format(page)
    print("第{0}页".format(page))
    print(url)
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'
        ,
        'Cookie': 'JSSearchModel=0; LastCity%5Fid=653; LastCity=%e6%9d%ad%e5%b7%9e; LastJobTag=%e4%ba%94%e9%99%a9%e4%b8%80%e9%87%91%7c%e8%8a%82%e6%97%a5%e7%a6%8f%e5%88%a9%7c%e5%b8%a6%e8%96%aa%e5%b9%b4%e5%81%87%7c%e7%bb%a9%e6%95%88%e5%a5%96%e9%87%91%7c%e9%a4%90%e8%a1%a5%7c%e5%ae%9a%e6%9c%9f%e4%bd%93%e6%a3%80%7c%e5%91%98%e5%b7%a5%e6%97%85%e6%b8%b8%7c%e5%bc%b9%e6%80%a7%e5%b7%a5%e4%bd%9c%7c%e5%b9%b4%e5%ba%95%e5%8f%8c%e8%96%aa%7c%e4%ba%a4%e9%80%9a%e8%a1%a5%e5%8a%a9%7c%e9%80%9a%e8%ae%af%e8%a1%a5%e8%b4%b4%7c%e9%ab%98%e6%b8%a9%e8%a1%a5%e8%b4%b4%7c%e8%82%a1%e7%a5%a8%e6%9c%9f%e6%9d%83%7c%e5%85%a8%e5%8b%a4%e5%a5%96%7c%e5%8a%a0%e7%8f%ad%e8%a1%a5%e5%8a%a9%7c%e8%a1%a5%e5%85%85%e5%8c%bb%e7%96%97%e4%bf%9d%e9%99%a9%7c%e5%85%8d%e8%b4%b9%e7%8f%ad%e8%bd%a6%7c%e5%b9%b4%e7%bb%88%e5%88%86%e7%ba%a2%7c%e5%8c%85%e5%90%83%7c%e6%88%bf%e8%a1%a5%7c%e5%8c%85%e4%bd%8f%7c%e9%87%87%e6%9a%96%e8%a1%a5%e8%b4%b4; LastSearchHistory=%7b%22Id%22%3a%22189ddb74-f21f-45b7-aae7-f4b2669b25a1%22%2c%22Name%22%3a%22python+%2b+%e6%9d%ad%e5%b7%9e%22%2c%22SearchUrl%22%3a%22http%3a%2f%2fsou.zhaopin.com%2fjobs%2fsearchresult.ashx%3fjl%3d%25e6%259d%25ad%25e5%25b7%259e%26kw%3dpython%26isadv%3d0%26sg%3db19e7c6f12a348359a72d76356038a60%26p%3d4%22%2c%22SaveTime%22%3a%22%5c%2fDate(1498753078650%2b0800)%5c%2f%22%7d; SubscibeCaptcha=6A68665FC75D4E6F42128088820FE28E; urlfrom=121126445; urlfrom2=121126445; adfcid=none; adfcid2=none; adfbid=0; adfbid2=0; dywez=95841923.1498753080.1.1.dywecsr=(direct)|dyweccn=(direct)|dywecmd=(none)|dywectr=undefined; dywea=95841923.196580863499777400.1498753080.1498753080.1498753080.1; dywec=95841923; dyweb=95841923.1.10.1498753080'
    }
    wbdata = requests.get(url, headers=header).text
    soup = BeautifulSoup(wbdata, 'lxml')

    job_name = soup.select("table.newlist > tr > td.zwmc > div > a")
    salarys = soup.select("table.newlist > tr > td.zwyx")
    locations = soup.select("table.newlist > tr > td.gzdd")
    times = soup.select("table.newlist > tr > td.gxsj > span")
    print('times  '+str(times))
    for name, salary, location, time in zip(job_name, salarys, locations, times):
        data = {
            'name': name.get_text(),
            'salary': salary.get_text(),
            'location': location.get_text(),
            'time': time.get_text(),
        }
        print(data)
        # ws.cell(row=page+1, column=1).value =name.get_text()
        # ws.cell(row=page+1, column=2).value = salary.get_text()
        # ws.cell(row=page+1, column=3).value = location.get_text()
        # ws.cell(row=page+1, column=4).value = time.get_text()

if __name__ == '__main__':
    pool = Pool(processes=6)
    pool.map_async(get_zhaopin, range(1, 91))
    wb.save('智联招聘电气' + '.xlsx')
    pool.close()
    pool.join()

    # for i in range(1,3):
    #     get_zhaopin(1)
    # wb.save('智联招聘的电气' + '.xlsx')
    print("-------------华丽的分割线-------------")

else:
     print("I am being imported from another module")
print('Done')